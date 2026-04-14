import reflex as rx
import asyncio
import csv
import io
from datetime import datetime, timezone, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill
from app.states.dashboard_state import DashboardState, ADVISOR_TARGETS
from app.services.hubspot_service import STAGES_VENTAS, STAGES_SOCIOS
import logging


class ReportState(DashboardState):
    report_type: str = "executive"
    report_schedule: str = "none"
    report_generating: bool = False
    report_generated: bool = False
    report_history: list[dict] = []
    temporal_period: str = "month"
    report_advisor_filter: str = "all"
    date_preset: str = "all"
    custom_start: str = ""
    custom_end: str = ""
    date_field: str = "createdate"
    compare_enabled: bool = False
    compare_mode: str = "previous"

    @rx.event
    def set_date_preset(self, preset: str):
        self.date_preset = preset

    @rx.event
    def set_custom_start(self, val: str):
        self.custom_start = val
        if val:
            self.date_preset = "custom"

    @rx.event
    def set_custom_end(self, val: str):
        self.custom_end = val
        if val:
            self.date_preset = "custom"

    @rx.event
    def set_date_field(self, field: str):
        self.date_field = field

    @rx.event
    def toggle_compare(self, val: bool):
        self.compare_enabled = val

    @rx.event
    def set_compare_mode(self, mode: str):
        self.compare_mode = mode

    @rx.event
    def set_report_type(self, val: str):
        self.report_type = val
        self.report_generated = False

    @rx.event
    def set_report_schedule(self, val: str):
        self.report_schedule = val

    @rx.event
    def set_temporal_period(self, val: str):
        self.temporal_period = val

    @rx.event
    def set_report_advisor_filter(self, val: str):
        self.report_advisor_filter = val

    @rx.event
    async def generate_report(self):
        self.report_generating = True
        yield
        await asyncio.sleep(1.5)
        self.report_generating = False
        self.report_generated = True
        now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
        self.report_history = [
            {
                "timestamp": now_str,
                "type": self.report_type,
                "filters": f"Pipeline: {self.selected_pipeline}, Period: {self.active_date_label}",
            }
        ] + self.report_history[:9]
        yield

    @rx.event
    def clear_history(self):
        self.report_history = []

    def _get_date_range(self) -> tuple[datetime | None, datetime | None]:
        now = datetime.now(timezone.utc)
        if self.date_preset == "all":
            return (None, None)
        elif self.date_preset == "today":
            start = now.replace(hour=0, minute=0, second=0, microsecond=0)
            end = now.replace(hour=23, minute=59, second=59, microsecond=999999)
            return (start, end)
        elif self.date_preset == "7d":
            return (now - timedelta(days=7), now)
        elif self.date_preset == "this_week":
            start = (now - timedelta(days=now.weekday())).replace(
                hour=0, minute=0, second=0, microsecond=0
            )
            return (start, now)
        elif self.date_preset == "last_week":
            end = (now - timedelta(days=now.weekday() + 1)).replace(
                hour=23, minute=59, second=59, microsecond=999999
            )
            start = (end - timedelta(days=6)).replace(
                hour=0, minute=0, second=0, microsecond=0
            )
            return (start, end)
        elif self.date_preset == "this_month":
            start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            return (start, now)
        elif self.date_preset == "last_month":
            end = (now.replace(day=1) - timedelta(days=1)).replace(
                hour=23, minute=59, second=59, microsecond=999999
            )
            start = end.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            return (start, end)
        elif self.date_preset == "this_quarter":
            q = (now.month - 1) // 3
            start = now.replace(
                month=q * 3 + 1, day=1, hour=0, minute=0, second=0, microsecond=0
            )
            return (start, now)
        elif self.date_preset == "last_quarter":
            q = (now.month - 1) // 3
            if q == 0:
                q = 3
                year = now.year - 1
            else:
                q -= 1
                year = now.year
            start = now.replace(
                year=year,
                month=q * 3 + 1,
                day=1,
                hour=0,
                minute=0,
                second=0,
                microsecond=0,
            )
            if q == 3:
                end = now.replace(
                    year=year,
                    month=12,
                    day=31,
                    hour=23,
                    minute=59,
                    second=59,
                    microsecond=999999,
                )
            else:
                end = (
                    now.replace(year=year, month=(q + 1) * 3 + 1, day=1)
                    - timedelta(days=1)
                ).replace(hour=23, minute=59, second=59, microsecond=999999)
            return (start, end)
        elif self.date_preset == "6m":
            return (now - timedelta(days=180), now)
        elif self.date_preset in ["this_year", "ytd"]:
            start = now.replace(
                month=1, day=1, hour=0, minute=0, second=0, microsecond=0
            )
            return (start, now)
        elif self.date_preset == "12m":
            return (now - timedelta(days=365), now)
        elif self.date_preset == "last_year":
            start = now.replace(
                year=now.year - 1,
                month=1,
                day=1,
                hour=0,
                minute=0,
                second=0,
                microsecond=0,
            )
            end = now.replace(
                year=now.year - 1,
                month=12,
                day=31,
                hour=23,
                minute=59,
                second=59,
                microsecond=999999,
            )
            return (start, end)
        elif self.date_preset == "custom" and self.custom_start and self.custom_end:
            try:
                start = datetime.fromisoformat(self.custom_start).replace(
                    tzinfo=timezone.utc
                )
                end = datetime.fromisoformat(self.custom_end).replace(
                    hour=23, minute=59, second=59, tzinfo=timezone.utc
                )
                return (start, end)
            except:
                logging.exception("Unexpected error")
        return (None, None)

    def _get_comparison_range(self) -> tuple[datetime | None, datetime | None]:
        start, end = self._get_date_range()
        if not start or not end:
            return (None, None)
        if self.compare_mode == "yoy":
            return (start.replace(year=start.year - 1), end.replace(year=end.year - 1))
        else:
            delta = end - start
            return (start - delta - timedelta(seconds=1), start - timedelta(seconds=1))

    @rx.var
    def active_date_label(self) -> str:
        labels = {
            "all": "Todo el tiempo",
            "today": "Hoy",
            "7d": "Últimos 7 días",
            "this_week": "Semana actual",
            "last_week": "Semana pasada",
            "this_month": "Mes actual",
            "last_month": "Mes anterior",
            "this_quarter": "Quarter actual",
            "last_quarter": "Quarter anterior",
            "6m": "Últimos 6 meses",
            "this_year": "Año actual",
            "12m": "Últimos 12 meses",
            "last_year": "Año anterior",
            "ytd": "Year to Date",
        }
        if self.date_preset == "custom":
            return f"{self.custom_start} → {self.custom_end}"
        return labels.get(self.date_preset, "Todo el tiempo")

    @rx.var
    def date_field_label(self) -> str:
        labels = {
            "createdate": "Fecha de creación",
            "closedate": "Fecha de cierre",
            "fecha_de_1er_cierre": "Fecha 1er cierre",
            "hs_lastmodifieddate": "Última actividad",
        }
        return labels.get(self.date_field, "Fecha")

    @rx.var
    def date_filtered_deals(self) -> list[dict]:
        start, end = self._get_date_range()
        if not start or not end:
            return self.filtered_deals
        filtered = []
        for d in self.filtered_deals:
            dt_str = d.get(self.date_field)
            if dt_str:
                try:
                    dt = datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
                    if dt.tzinfo is None:
                        dt = dt.replace(tzinfo=timezone.utc)
                    if start <= dt <= end:
                        filtered.append(d)
                except:
                    logging.exception("Unexpected date parsing error")
        return filtered

    @rx.var
    def comparison_deals(self) -> list[dict]:
        if not self.compare_enabled:
            return []
        start, end = self._get_comparison_range()
        if not start or not end:
            return []
        filtered = []
        for d in self.filtered_deals:
            dt_str = d.get(self.date_field)
            if dt_str:
                try:
                    dt = datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
                    if dt.tzinfo is None:
                        dt = dt.replace(tzinfo=timezone.utc)
                    if start <= dt <= end:
                        filtered.append(d)
                except:
                    logging.exception("Unexpected date parsing error")
        return filtered

    @rx.var
    def comparison_stats(self) -> dict:
        if not self.compare_enabled:
            return {}
        curr = self.date_filtered_deals
        prev = self.comparison_deals
        c_deals = len(curr)
        p_deals = len(prev)
        c_kwh = sum([float(d.get("capacidad__kwh___clonada_") or 0) for d in curr])
        p_kwh = sum([float(d.get("capacidad__kwh___clonada_") or 0) for d in prev])
        c_usd = sum([float(d.get("amount") or 0) for d in curr])
        p_usd = sum([float(d.get("amount") or 0) for d in prev])
        c_won = sum(
            [1 for d in curr if d.get("dealstage") in ["closedwon", "1182079922"]]
        )
        p_won = sum(
            [1 for d in prev if d.get("dealstage") in ["closedwon", "1182079922"]]
        )
        return {
            "current_deals": c_deals,
            "previous_deals": p_deals,
            "deals_delta": f"{('+' if c_deals - p_deals >= 0 else '')}{c_deals - p_deals}",
            "deals_growth_pct": f"{((c_deals - p_deals) / p_deals * 100 if p_deals > 0 else 0):.1f}%",
            "current_kwh": c_kwh,
            "previous_kwh": p_kwh,
            "kwh_delta": f"{('+' if c_kwh - p_kwh >= 0 else '')}{c_kwh - p_kwh:,.0f}",
            "kwh_growth_pct": f"{((c_kwh - p_kwh) / p_kwh * 100 if p_kwh > 0 else 0):.1f}%",
            "current_usd": c_usd,
            "previous_usd": p_usd,
            "usd_delta": f"{('+$' if c_usd - p_usd >= 0 else '-$')}{abs(c_usd - p_usd):,.0f}",
            "usd_growth_pct": f"{((c_usd - p_usd) / p_usd * 100 if p_usd > 0 else 0):.1f}%",
            "current_won": c_won,
            "previous_won": p_won,
            "won_delta": f"{('+' if c_won - p_won >= 0 else '')}{c_won - p_won}",
            "won_growth_pct": f"{((c_won - p_won) / p_won * 100 if p_won > 0 else 0):.1f}%",
            "current_label": self.active_date_label,
            "previous_label": "Mismo periodo anterior"
            if self.compare_mode == "previous"
            else "Año anterior",
        }

    @rx.var
    def report_insights(self) -> list[dict]:
        insights = []
        adv_forecast = self.advisor_forecast_data
        proj = sum((a.get("projected", 0) for a in adv_forecast))
        if proj >= self.total_target_q and self.total_target_q > 0:
            insights.append(
                {
                    "type": "forecast",
                    "icon": "trending-up",
                    "title": "Forecast Trimestral",
                    "description": f"El equipo proyecta alcanzar la meta. ({proj:,.0f} / {self.total_target_q:,.0f} kWh)",
                    "severity": "info",
                    "color": "#10B981",
                }
            )
        else:
            insights.append(
                {
                    "type": "forecast",
                    "icon": "trending-down",
                    "title": "Riesgo de Forecast",
                    "description": f"El equipo proyecta quedar corto por {self.total_target_q - proj:,.0f} kWh.",
                    "severity": "warning",
                    "color": "#F59E0B",
                }
            )
        stale_count = 0
        now = datetime.now(timezone.utc)
        for d in self.date_filtered_deals:
            stage = d.get("dealstage")
            if stage not in ["closedlost", "closedwon", "1182079923", "1182079922"]:
                lm_str = d.get("hs_lastmodifieddate")
                if lm_str:
                    try:
                        dt = datetime.fromisoformat(lm_str.replace("Z", "+00:00"))
                        if dt.tzinfo is None:
                            dt = dt.replace(tzinfo=timezone.utc)
                        if (now - dt).days >= 30:
                            stale_count += 1
                    except:
                        logging.exception("Unexpected error")
        if stale_count > 0:
            insights.append(
                {
                    "type": "alert",
                    "icon": "alert-triangle",
                    "title": "Negocios Estancados",
                    "description": f"Existen {stale_count} negocios activos sin actividad en más de 30 días.",
                    "severity": "warning",
                    "color": "#F59E0B",
                }
            )
        stats = self.advisor_stats
        if len(stats) >= 3:
            top_names = ", ".join([s["name"][:10] for s in stats[:3]])
            insights.append(
                {
                    "type": "star",
                    "icon": "star",
                    "title": "Top Performers",
                    "description": f"Líderes en kWh: {top_names}",
                    "severity": "info",
                    "color": "#22D3EE",
                }
            )
        if self.total_target_q > 0:
            gap_pct = (
                self.total_target_q - self.total_kwh_cerrados_q
            ) / self.total_target_q
            if gap_pct > 0.3:
                insights.append(
                    {
                        "type": "gap",
                        "icon": "target",
                        "title": "Brecha de Meta Alta",
                        "description": f"Falta el {gap_pct * 100:.1f}% de la meta trimestral.",
                        "severity": "critical",
                        "color": "#EF4444",
                    }
                )
        overdue_count = 0
        for d in self.date_filtered_deals:
            stage = d.get("dealstage")
            if stage not in ["closedlost", "closedwon", "1182079923", "1182079922"]:
                cd_str = d.get("closedate")
                if cd_str:
                    try:
                        dt = datetime.fromisoformat(cd_str.replace("Z", "+00:00"))
                        if dt.tzinfo is None:
                            dt = dt.replace(tzinfo=timezone.utc)
                        if dt < now:
                            overdue_count += 1
                    except:
                        logging.exception("Unexpected error")
        if overdue_count > 5:
            insights.append(
                {
                    "type": "overdue",
                    "icon": "clock",
                    "title": "Cierres Atrasados",
                    "description": f"{overdue_count} negocios han pasado su fecha de cierre estimada.",
                    "severity": "critical",
                    "color": "#EF4444",
                }
            )
        zero_count = sum((1 for s in stats if s.get("kwh_cerrados", 0) == 0))
        if zero_count > 0:
            insights.append(
                {
                    "type": "anomaly",
                    "icon": "user-x",
                    "title": "Asesores sin Cierres",
                    "description": f"{zero_count} asesores no tienen kWh cerrados en el trimestre.",
                    "severity": "warning",
                    "color": "#F59E0B",
                }
            )
        return insights

    @rx.var
    def executive_report_data(
        self,
    ) -> dict[
        str,
        str | int | float | bool | list[dict[str, str | int]] | dict[str, str | bool],
    ]:
        adv_forecast = self.advisor_forecast_data
        proj = sum((a.get("projected", 0) for a in adv_forecast))
        will_meet = bool(proj >= self.total_target_q)
        insights_data = self.report_insights
        deals_source = self.date_filtered_deals
        total_deals = len(deals_source)
        active_deals = 0
        convenio_firmado = 0
        cerrados_perdidos = 0
        total_usd = 0.0
        total_kwh = 0.0
        kwh_cerrados = 0.0
        kwh_activos = 0.0
        kwh_perdidos = 0.0
        for d in deals_source:
            stage = d.get("dealstage")
            pipe = d.get("pipeline")
            if pipe == "default" and stage not in ["closedlost", "closedwon"]:
                active_deals += 1
            elif pipe == "803674731" and stage != "1182079923":
                active_deals += 1
            if stage in ["closedwon", "1182079922"]:
                convenio_firmado += 1
            elif stage in ["closedlost", "1182079923"]:
                cerrados_perdidos += 1
            amt = d.get("amount")
            if amt:
                try:
                    total_usd += float(amt)
                except:
                    logging.exception("Unexpected error")
            k_val = d.get("capacidad__kwh___clonada_")
            if k_val:
                try:
                    k_flt = float(k_val)
                    total_kwh += k_flt
                    if stage in ["closedwon", "1182079922"]:
                        kwh_cerrados += k_flt
                    elif stage in ["closedlost", "1182079923"]:
                        kwh_perdidos += k_flt
                    else:
                        kwh_activos += k_flt
                except:
                    logging.exception("Unexpected error")
        total_usd_fmt = self._format_compact_currency(total_usd)
        total_kwh_fmt = f"{total_kwh:,.0f}"
        return {
            "total_deals": total_deals,
            "active_deals": active_deals,
            "convenio_firmado": convenio_firmado,
            "cerrados_perdidos": cerrados_perdidos,
            "total_usd": total_usd_fmt,
            "total_kwh": total_kwh_fmt,
            "kwh_cerrados": kwh_cerrados,
            "kwh_activos": kwh_activos,
            "kwh_perdidos": kwh_perdidos,
            "overall_pct_q": float(self.overall_pct_q),
            "current_quarter_label": str(self.current_quarter_label),
            "total_kwh_cerrados_q_fmt": str(self.total_kwh_cerrados_q_fmt),
            "total_target_q_fmt": str(self.total_target_q_fmt),
            "top_advisors": [
                {
                    "name": str(a["name"]),
                    "total": int(a["total"]),
                    "usd_fmt": str(a["usd_fmt"]),
                    "kwh_fmt": str(a["kwh_fmt"]),
                    "won_count": int(a["convenio"]),
                    "rank": int(i + 1),
                }
                for i, a in enumerate(self.advisor_stats[:5])
            ],
            "forecast_summary": {"projected": f"{proj:,.0f}", "will_meet": will_meet},
            "risks": [
                {
                    "description": str(i.get("description", "")),
                    "title": str(i.get("title", "")),
                    "severity": str(i.get("severity", "")),
                }
                for i in insights_data
                if i["severity"] in ["critical", "warning"]
            ],
            "insights": [
                {
                    "description": str(i.get("description", "")),
                    "title": str(i.get("title", "")),
                    "severity": str(i.get("severity", "")),
                }
                for i in insights_data
                if i["severity"] == "info"
            ],
        }

    @rx.var
    def pipeline_report_data(self) -> dict:
        now = datetime.now(timezone.utc)
        stale_deals = []
        ventas_counts = {
            "1230967589": 0,
            "appointmentscheduled": 0,
            "qualifiedtobuy": 0,
            "1329256222": 0,
            "contractsent": 0,
            "1212783978": 0,
            "1212794259": 0,
            "closedwon": 0,
            "1233929775": 0,
            "1232935533": 0,
            "1232929105": 0,
            "closedlost": 0,
        }
        ventas_total = 0
        socios_counts = {
            "1182079917": 0,
            "1185807635": 0,
            "1329255962": 0,
            "1338977193": 0,
            "1182079922": 0,
            "1182079923": 0,
        }
        socios_total = 0
        for d in self.date_filtered_deals:
            stage = d.get("dealstage")
            pipe = d.get("pipeline")
            if pipe == "default" and stage in ventas_counts:
                ventas_counts[stage] += 1
                ventas_total += 1
            elif pipe == "803674731" and stage in socios_counts:
                socios_counts[stage] += 1
                socios_total += 1
            if stage not in ["closedlost", "closedwon", "1182079923", "1182079922"]:
                lm_str = d.get("hs_lastmodifieddate")
                if lm_str:
                    try:
                        dt = datetime.fromisoformat(lm_str.replace("Z", "+00:00"))
                        if dt.tzinfo is None:
                            dt = dt.replace(tzinfo=timezone.utc)
                        days_stale = (now - dt).days
                        if days_stale >= 30:
                            owner = self.owners.get(
                                d.get("hubspot_owner_id", ""), "Sin asignar"
                            )
                            stale_deals.append(
                                {
                                    "dealname": d.get("dealname", "Sin nombre")[:30],
                                    "owner": owner,
                                    "days_stale": days_stale,
                                    "stage": stage or "Sin etapa",
                                }
                            )
                    except:
                        logging.exception("Unexpected error")
        stale_deals.sort(key=lambda x: x["days_stale"], reverse=True)
        ventas_res = []
        for st, c in ventas_counts.items():
            pct = f"{c / ventas_total * 100:.1f}%" if ventas_total > 0 else "0%"
            ventas_res.append(
                {"stage_name": STAGES_VENTAS.get(st, st), "count": c, "percentage": pct}
            )
        socios_res = []
        for st, c in socios_counts.items():
            pct = f"{c / socios_total * 100:.1f}%" if socios_total > 0 else "0%"
            socios_res.append(
                {"stage_name": STAGES_SOCIOS.get(st, st), "count": c, "percentage": pct}
            )
        return {
            "ventas_stages": ventas_res,
            "socios_stages": socios_res,
            "aging_data": [],
            "stale_deals": stale_deals[:10],
            "stage_flow": self.stage_flow_data,
            "advancement_rates": [],
        }

    @rx.var
    def advisor_report_data(self) -> list[dict]:
        return self.advisor_quota_data

    @rx.var
    def deal_explorer_data(self) -> list[dict]:
        deals = []
        now = datetime.now(timezone.utc)
        for d in self.date_filtered_deals[:50]:
            kwh = d.get("capacidad__kwh___clonada_")
            amount = d.get("amount")
            stage = d.get("dealstage")
            days_in_stage = 0
            lm_str = d.get("hs_v2_date_entered_current_stage") or d.get(
                "hs_lastmodifieddate"
            )
            if lm_str:
                try:
                    dt = datetime.fromisoformat(lm_str.replace("Z", "+00:00"))
                    if dt.tzinfo is None:
                        dt = dt.replace(tzinfo=timezone.utc)
                    days_in_stage = (now - dt).days
                except:
                    logging.exception("Unexpected error")
            risk_score = "Bajo"
            if stage not in ["closedlost", "closedwon", "1182079923", "1182079922"]:
                if days_in_stage > 60:
                    risk_score = "Alto"
                elif days_in_stage > 30:
                    risk_score = "Medio"
                cd_str = d.get("closedate")
                if cd_str:
                    try:
                        dt = datetime.fromisoformat(cd_str.replace("Z", "+00:00"))
                        if dt.tzinfo is None:
                            dt = dt.replace(tzinfo=timezone.utc)
                        if dt < now:
                            risk_score = "Alto"
                    except:
                        logging.exception("Unexpected error")
            deals.append(
                {
                    "dealname": d.get("dealname", "Sin nombre")[:30],
                    "owner_name": self.owners.get(
                        d.get("hubspot_owner_id"), "Sin asignar"
                    ),
                    "pipeline_name": "Ventas"
                    if d.get("pipeline") == "default"
                    else "Socios"
                    if d.get("pipeline") == "803674731"
                    else "Otro",
                    "stage_name": stage or "Sin etapa",
                    "amount_fmt": f"${float(amount):,.0f}" if amount else "$0",
                    "kwh_fmt": f"{float(kwh):,.0f}" if kwh else "0",
                    "days_in_stage": days_in_stage,
                    "socio": d.get("referido_por_socio") or "—",
                    "risk_score": risk_score,
                    "created": str(d.get("createdate", ""))[:10],
                    "close_date": str(d.get("closedate", ""))[:10],
                }
            )
        return deals

    @rx.var
    def temporal_report_data(self) -> dict:
        counts = {}
        for d in self.date_filtered_deals:
            dt_str = d.get("createdate")
            if dt_str:
                try:
                    dt = datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
                    if self.temporal_period == "year":
                        key = str(dt.year)
                    elif self.temporal_period == "quarter":
                        q = (dt.month - 1) // 3 + 1
                        key = f"Q{q} {dt.year}"
                    else:
                        key = dt.strftime("%Y-%m")
                    if key not in counts:
                        counts[key] = {"deals": 0, "kwh": 0.0, "usd": 0.0, "won": 0}
                    counts[key]["deals"] += 1
                    k = d.get("capacidad__kwh___clonada_")
                    if k:
                        counts[key]["kwh"] += float(k)
                    u = d.get("amount")
                    if u:
                        counts[key]["usd"] += float(u)
                    st = d.get("dealstage")
                    if st in ["closedwon", "1182079922"]:
                        counts[key]["won"] += 1
                except:
                    logging.exception("Unexpected error")
        periods = []
        for k in sorted(counts.keys())[-12:]:
            c = counts[k]
            periods.append(
                {
                    "label": k,
                    "deals": c["deals"],
                    "kwh_fmt": f"{c['kwh']:,.0f}",
                    "usd_fmt": f"${c['usd']:,.0f}",
                    "won": c["won"],
                }
            )
        return {
            "periods": periods,
            "mom_comparison": {},
            "qoq_comparison": {},
            "yoy_comparison": {},
        }

    @rx.var
    def report_meta(self) -> dict:
        return {
            "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
            "generated_by": "Quartux Admin",
            "filters_applied": f"Pipeline: {self.selected_pipeline}",
            "last_sync": self.last_sync_timestamp,
            "report_type_label": self.report_type.capitalize(),
            "deal_count": len(self.date_filtered_deals),
            "date_range": self.active_date_label,
            "date_field": self.date_field_label,
        }

    @rx.event
    def export_csv(self):
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Período", self.active_date_label])
        writer.writerow(["Campo de fecha", self.date_field_label])
        writer.writerow(
            ["Generado", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")]
        )
        writer.writerow([])
        writer.writerow(["Métrica", "Valor"])
        writer.writerow(["Total Negocios", len(self.date_filtered_deals)])
        data = output.getvalue()
        return rx.download(data=data, filename=f"reporte_{self.report_type}.csv")

    @rx.event
    def export_excel(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte"
        ws.append(["Período", self.active_date_label])
        ws.append(["Campo de fecha", self.date_field_label])
        ws.append(
            ["Generado", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")]
        )
        ws.append([])
        header_fill = PatternFill(
            start_color="DC2626", end_color="DC2626", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)
        headers = ["Métrica", "Valor"]
        ws.append(headers)
        for col_num in range(1, 3):
            cell = ws.cell(row=5, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
        ws.append(["Total Negocios", len(self.date_filtered_deals)])
        output = io.BytesIO()
        wb.save(output)
        return rx.download(
            data=output.getvalue(), filename=f"reporte_{self.report_type}.xlsx"
        )